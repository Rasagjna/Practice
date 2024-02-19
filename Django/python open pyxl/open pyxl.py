from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
# wb=load_workbook('Book1.xlsx')
# ws=wb.active
# # print(ws['A1'].value)
# # ws['A2'].value="Test"
# # wb.save('Books1.xlsx')
# print(wb.sheetnames)
# wb=Workbook()
# ws=wb.active
# ws.title="Data"
# ws.append(['Tim','Is','Great','!'])
# ws.append(['Tim','Is','Great','!'])
# ws.append(['Tim','Is','Great','!'])
# ws.append(['Tim','Is','Great','!'])
# ws.append(['Tim','Is','Great','!'])
# ws.append(['Tim','Is'])
# wb.save('tim.xlsx')


wb=load_workbook('tim.xlsx')
ws=wb.active
for row in range(1,11):
    for col in range(1,5):
        char=get_column_letter(col)
        print(ws[char+str(row)])
        
    
    